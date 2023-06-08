#!/usr/bin/env python

import json
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os
import sys
import yaml


def usage():
    print("Converts a directory of schemas to an excel spreadsheet template.")
    print("")
    print("Usage:")
    print("    " + " ".join([sys.argv[0],
                             "<xlsx_file>",
                             "[example_data_dir]"]))
    print("")


class ExcelTemplateGenerator(object):

    def __init__(self):
        self.settings = {
            "schema_directory": "schemas",
            "schema_order": [],
            "column_offset": 1,
            "row_offset": 4,
            "object_width": 70,
            "column_width": 25,
            "comment_width": 600,
            "comment_height": 300,
            "required_color": "FFFFDD",
            "normal_color": "FFFFFF",
            "advanced_color": "EEEEEE",
            "advanced_text_color": "888888",
            "row_label_color": "AAAAAA",
            "border_color": "AAAAAA",
            "section_color": "8888FF",
            "section_text_color": "FFFFFF",
            "num_list_entries": 10,
            "default_fields_by_col": True
        }

    def write_workbook(self, output_file, example_dir=None):
        workbook = Workbook()

        workbook.remove(workbook.active)

        schema_names = self.find_schema_names()

        for schema_name in schema_names:
            self.generate_worksheet(workbook, schema_name, example_dir)

        # workbook.template = True
        workbook.save(output_file)

    def find_schema_names(self):
        schema_names = self.settings["schema_order"]
        for file_name in sorted(os.listdir(self.settings["schema_directory"])):
            if (file_name.endswith(".json")):
                schema_name = file_name[0:-5]
                if schema_name not in schema_names:
                    schema_names.append(schema_name)

        return schema_names

    def generate_worksheet(self, workbook, schema_name, example_dir=None):
        schema = self.read_schema(schema_name)
        example = None
        if example_dir is not None:
            example = self.read_example(example_dir, schema_name)

        worksheet = workbook.create_sheet(self.name_to_title(schema_name))

        self.generate_title(worksheet, schema)

        fields_by_col = self.settings["default_fields_by_col"]
        if "fieldsByCol" in schema:
            fields_by_col = schema["fieldsByCol"]

        if (schema["type"] == "array" and fields_by_col):
            self.generate_schema_list(worksheet, schema, example)
        else:
            self.generate_schema_object(worksheet, schema, example)

        return worksheet

    def generate_title(self, worksheet, schema):
        worksheet["A1"] = schema["title"]
        worksheet["A1"].style = "Title"
        worksheet["A2"] = schema["description"]
        worksheet["A2"].style = "Headline 3"

    def generate_schema_object(self, worksheet, schema, example=None):
        row_offset = self.settings["row_offset"]
        col_offset = self.settings["column_offset"]

        is_list = schema["type"] == "array"

        if is_list:
            properties = schema["items"]["properties"]
            num_rows = self.get_num_rows(schema, example)
            list_name = self.get_list_name(schema)
        else:
            properties = schema["properties"]
            num_rows = 1

        i = 0
        for name, field in sorted(iter(properties.items()),
                                  key=lambda v: v[1]["propertyOrder"]):
            field["name"] = name
            if is_list:
                field["required"] = (
                    "required" in schema["items"] and
                    name in schema["items"]["required"])
            else:
                field["required"] = (
                    "required" in schema and name in schema["required"])

            if "sectionBegin" in field:
                cell_start = worksheet.cell(row=row_offset + i,
                                            column=col_offset)
                cell_end = worksheet.cell(row=row_offset + i,
                                          column=col_offset + num_rows)
                self.write_section_cells(worksheet,
                                         field["sectionBegin"],
                                         cell_start, cell_end)
                i += 1

            self.write_label_cell(worksheet, field, i + row_offset,
                                  col_offset)

            for j in range(num_rows):

                if is_list:
                    cell = worksheet.cell(row=1,
                                          column=j + col_offset + 1)
                    col = worksheet.column_dimensions[cell.column_letter]
                    col.width = self.settings["column_width"]
                    value = self.get_example_row_value(example, name,
                                                       list_name, j)
                else:
                    value = None
                    if example is not None and name in example:
                        value = example[name]

                self.write_field_cell(worksheet, field, i + row_offset,
                                      col_offset + j + 1, value)
                self.add_data_validation(worksheet, field, i + row_offset,
                                         col_offset + j + 1)
            i += 1

        col = worksheet.column_dimensions['A']
        col.width = self.settings["object_width"]
        if not is_list:
            col = worksheet.column_dimensions['B']
            col.width = self.settings["object_width"]

    def generate_schema_list(self, worksheet, schema, example=None):
        row_offset = self.settings["row_offset"]
        col_offset = self.settings["column_offset"]
        properties = schema["items"]["properties"]
        num_rows = self.get_num_rows(schema, example)
        list_name = self.get_list_name(schema)

        self.write_row_sections(worksheet, properties)
        row_offset += 1

        i = 0
        for name, field in sorted(iter(properties.items()),
                                  key=lambda v: v[1]["propertyOrder"]):
            field["name"] = name
            field["required"] = ("required" in schema["items"] and
                                 name in schema["items"]["required"])

            self.write_label_cell(worksheet, field, row_offset,
                                  i + col_offset, row_label=True)

            cell = worksheet.cell(row=row_offset, column=i + col_offset)
            col = worksheet.column_dimensions[cell.column_letter]
            col.width = self.settings["column_width"]

            for j in range(num_rows):
                value = self.get_example_row_value(example, name, list_name, j)

                self.write_field_cell(worksheet, field, row_offset + j + 1,
                                      i + col_offset, value)
                self.add_data_validation(worksheet, field, row_offset + j + 1,
                                         i + col_offset)
            i += 1

    def write_row_sections(self, worksheet, properties):
        row_offset = self.settings["row_offset"]
        col_offset = self.settings["column_offset"]

        cell_start = None
        section_name = "(missing)"
        i = 0
        for name, field in sorted(iter(properties.items()),
                                  key=lambda v: v[1]["propertyOrder"]):
            if "sectionBegin" in field:
                cell_start = worksheet.cell(row=row_offset,
                                            column=i + col_offset)
                section_name = field["sectionBegin"]

            if "sectionEnd" in field and cell_start is not None:
                cell_end = worksheet.cell(row=row_offset,
                                          column=i + col_offset)

                self.write_section_cells(worksheet, section_name,
                                         cell_start, cell_end,
                                         border=True)

            i += 1

    def get_num_rows(self, schema, example):
        num_rows = self.settings["num_list_entries"]
        list_name = self.get_list_name(schema)
        if example is not None and list_name in example:
            num_rows = len(example[list_name])
        if "numFormEntries" in schema:
            num_rows = schema["numFormEntries"]

        return num_rows

    def get_example_row_value(self, example, name, list_name, index):
        value = None
        if type(example) == list:
            example = {list_name: example}
        if (example is not None and list_name in example and
                index < len(example[list_name])):
            example_row = example[list_name][index]
            if name in example_row:
                value = example_row[name]

        return value

    def write_label_cell(self, worksheet, field, row, col, row_label=False):
        cell = worksheet.cell(row=row, column=col)
        cell.value = field["title"]
        default = ""
        if "default" in field:
            default = " [default: %s]" % field["default"]

        description = ""
        if "description" in field:
            description = field["description"]

        list_info = ""
        if "type" in field and field["type"] == "array":
            list_info = " (List items separated by comma.)"

        cell.comment = Comment(description + default + list_info,
                               field["name"])
        cell.comment.width = self.settings["comment_width"]
        cell.comment.height = self.settings["comment_height"]
        if row_label and "required" in field and field["required"]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid",
                                    fgColor=self.settings["row_label_color"])
        elif row_label:
            cell.fill = PatternFill("solid",
                                    fgColor=self.settings["row_label_color"])
        elif "required" in field and field["required"]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid",
                                    fgColor=self.settings["required_color"])
        elif "advanced" in field and field["advanced"]:
            cell.font = Font(color=self.settings["advanced_text_color"])
            cell.fill = PatternFill("solid",
                                    fgColor=self.settings["advanced_color"])
        else:
            cell.fill = PatternFill("solid",
                                    fgColor=self.settings["normal_color"])

    def write_field_cell(self, worksheet, field, row, col, value=None):
        cell = worksheet.cell(row=row, column=col)
        if value is not None:
            if type(value) == list:
                cell.value = ", ".join(value)
            else:
                cell.value = value
        thin = Side(border_style="thin", color=self.settings["border_color"])
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        if "required" in field and field["required"]:
            cell.fill = PatternFill("solid",
                                    fgColor=self.settings["required_color"])
        elif "advanced" in field and field["advanced"]:
            cell.fill = PatternFill("solid",
                                    fgColor=self.settings["advanced_color"])
        else:
            cell.fill = PatternFill("solid",
                                    fgColor=self.settings["normal_color"])

    def write_section_cells(self, worksheet, name, cell_start, cell_end,
                            border=False):
        worksheet.merge_cells(cell_start.coordinate + ":" +
                              cell_end.coordinate)

        cell_start.font = Font(color=self.settings["section_text_color"])
        cell_start.fill = PatternFill("solid",
                                      fgColor=self.settings["section_color"])
        cell_start.alignment = Alignment(horizontal="center",
                                         vertical="center")
        cell_start.value = name

        if border:
            thin = Side(border_style="thin",
                        color=self.settings["section_text_color"])
            cell_start.border = Border(left=thin, right=thin)

    def add_data_validation(self, worksheet, field, row, col):
        if "enum" in field:
            dv = DataValidation(type="list",
                                formula1='"%s"' % ",".join(field["enum"]),
                                allow_blank=True, errorStyle='warning')
            dv.error = "Your entry is not in the list, Change anyway?"
            dv.errorTitle = "Invalid Entry"
            dv.prompt = "Please select from the list"
            dv.promptTitle = "List Selection"
        elif "type" not in field:
            return
        elif field["type"] == "boolean":
            dv = DataValidation(type="list",
                                formula1='"true,false"',
                                allow_blank=True, errorStyle='warning')
            dv.error = "Your entry is not true or false, change anyway?"
            dv.errorTitle = "Invalid Entry"
            dv.prompt = "Please select true or false"
            dv.promptTitle = "True or False Selection"
        elif field["type"] == "integer":
            dv = DataValidation(type="whole",
                                allow_blank=True, errorStyle="warning")
            dv.error = "Your entry is not an integer, change anyway?"
            dv.errorTitle = "Invalid Entry"
            dv.prompt = "Please provide integer"
            dv.promptTitle = "Integer Selection"
        else:
            return

        worksheet.add_data_validation(dv)
        c = worksheet.cell(row=row, column=col)
        dv.add(c)

    def read_schema(self, schema_name):
        file_name = schema_name + ".json"
        file_path = os.path.join(self.settings["schema_directory"], file_name)
        with open(file_path, "r", encoding="utf-8") as f:
            schema_str = f.read()

        try:
            return json.loads(schema_str)
        except Exception as e:
            raise Exception("Could not parse schema: %s\n%s" % (
                file_name, str(e)))

    def name_to_title(self, name):
        return name[0].upper() + name[1:].replace("_", " ")

    def read_example(self, example_dir, schema_name):
        file_name = os.path.join(example_dir, schema_name + ".yml")
        if os.path.isfile(file_name):
            try:
                with open(file_name, "r", encoding="utf-8") as f:
                    return yaml.safe_load(f.read())
            except Exception as e:
                raise Exception("Could not parse example: %s\n%s" % (
                    file_name, str(e)))
        else:
            return None

    def get_list_name(self, schema):
        if "listName" in schema:
            list_name = schema["listName"]
        else:
            if "items" in schema and "title" in schema["items"]:
                list_name = (
                    schema["items"]["title"].lower().replace(" ", "_") + "s")
            else:
                list_name = schema["title"].lower().replace(" ", "_")

        return list_name


def main():
    if len(sys.argv) not in [2, 3]:
        usage()
        exit(1)

    xlsx_file = sys.argv[1]
    example_data_dir = None
    if len(sys.argv) == 3:
        example_data_dir = sys.argv[2]

    generator = ExcelTemplateGenerator()
    generator.settings["schema_order"] = ["deployment", "common", "upgrade",
                                          "vsds", "vscs", "vstats"]
    generator.settings["num_list_entries"] = 6
    generator.settings["default_fields_by_col"] = False

    generator.write_workbook(xlsx_file, example_data_dir)


if __name__ == '__main__':
    main()
