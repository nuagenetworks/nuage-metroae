#!/usr/bin/env python
import sys
import yaml
import getpass
import argparse
import os
from ansible.parsing.vault import VaultEditor, VaultSecret, is_encrypted
from openpyxl import load_workbook


class ExcelParseError(Exception):
    pass


class ExcelParser(object):

    def __init__(self):
        self.settings = {
            "column_offset": 1,
            "row_offset": 4,
            "row_sections_present": True,
            "use_list_name": False,
            "default_fields_by_col": True}

        self.errors = list()
        self.cell_positions = dict()

    def read_and_encrypt_credentials_excel_sheet(self, passcode, spreadsheet_path):
        with open('schemas/credentials.json') as credentials_schema:
            schema_data = yaml.load(credentials_schema.read().decode("utf-8"),
                                    Loader=yaml.Loader)
        props = schema_data['items']['properties']

        title_field_map = self.generate_title_field_map(props)
        fields_by_col = self.settings["default_fields_by_col"]
        workbook = load_workbook(spreadsheet_path)
        worksheet = workbook['Credentials']
        labels = self.read_labels(worksheet, title_field_map,
                                  fields_by_col=fields_by_col)

        entry_offset = 0
        do_not_encrypt_list = self.get_do_not_encrypt_list()
        while True:
            self.cell_positions.clear()
            entry = self.read_and_encrypt_data_entry(workbook, spreadsheet_path, worksheet,
                                                     labels, entry_offset, do_not_encrypt_list,
                                                     passcode, fields_by_col=fields_by_col)

            if entry != dict():
                entry_offset += 1
            else:
                break

    def generate_title_field_map(self, properties):
        title_field_map = dict()
        for name, field in iter(properties.items()):
            if "type" in field and field["type"] == "array":
                title_field_map[field["title"]] = "list:" + name
            else:
                title_field_map[field["title"]] = name

        return title_field_map

    def read_labels(self, worksheet, title_field_map, fields_by_col=False):
        labels = list()

        col = self.settings["column_offset"]
        row = self.settings["row_offset"]

        if self.settings["row_sections_present"] and fields_by_col:
            row += 1

        while True:
            cell = worksheet.cell(row=row, column=col)
            value = cell.value
            if fields_by_col:
                col += 1
            else:
                row += 1

            if value is not None:
                if value in title_field_map:
                    labels.append(title_field_map[value])
                else:
                    labels.append(None)
            else:
                break

        return labels

    def read_and_encrypt_data_entry(self, wb, file_path, worksheet, labels, entry_offset,
                                    do_not_encrypt_list, passcode, fields_by_col=False):
        entry = dict()

        col = self.settings["column_offset"]
        row = self.settings["row_offset"]

        if fields_by_col:
            row += entry_offset + 1
            if self.settings["row_sections_present"]:
                row += 1
        else:
            col += entry_offset + 1

        for label in labels:
            cell = worksheet.cell(row=row, column=col)
            value = cell.value
            if value is not None:
                if label is not None:
                    if label not in do_not_encrypt_list:
                        encrypted_value = self.encrypt_value(value, passcode)
                        cell.value = encrypted_value
                        wb.save(file_path)
                        entry[label] = encrypted_value
                    self.cell_positions[label] = cell.coordinate
                else:
                    self.record_error(cell.coordinate, "Data entry for unknown label")
            else:
                self.cell_positions[label] = cell.coordinate
            if fields_by_col:
                col += 1
            else:
                row += 1

        return entry

    def get_do_not_encrypt_list(self):
        with open('schemas/credentials.json') as credentials_schema:
            data = yaml.load(credentials_schema.read().decode("utf-8"),
                             Loader=yaml.Loader)
        props = data['items']['properties']
        do_not_encrypt_list = []
        for k, v in props.items():
            if ('encrypt' in v) and (not v['encrypt']):
                do_not_encrypt_list.append(k)

        return do_not_encrypt_list

    def encrypt_value(self, value, passcode):
        secret = VaultSecret(passcode)
        editor = VaultEditor()
        if not is_encrypted(value):
            vaultCode = editor.encrypt_bytes(value, secret)
        else:
            vaultCode = value
        encrypted_val = '!vault |\n' + (vaultCode)

        return encrypted_val

    def record_error(self, position, message):
        self.errors.append({"position": position,
                            "message": message})


def main():
    parser = argparse.ArgumentParser(
        description='Encrypt user credentials in Excel spreadsheet')
    parser.add_argument(
        "deployment_spreadsheet_path",
        nargs='?',
        help="Required path to excel deployment spreadsheet")
    args = parser.parse_args()

    print("ARGS: ", args.deployment_spreadsheet_path)

    if "METROAE_PASSWORD" in os.environ:
        passcode = os.environ["METROAE_PASSWORD"]
    else:
        try:
            print("This file will encrypt user credentials for MetroAE")
            print("All user comments and unsupported fields in the "
                  "credentials file will be lost")
            print("Press Ctrl-C to cancel")
            while True:
                passcode = getpass.getpass()
                confirm_passcode = getpass.getpass("Confirm passcode:")
                if passcode != confirm_passcode:
                    print("Passcodes do not match. Please reenter")
                else:
                    break
        except Exception:
            print("Error in getting passcode from command line")
            sys.exit()

    parser = ExcelParser()
    parser.settings["use_list_name"] = True
    parser.settings["default_fields_by_col"] = False

    try:
        parser.read_and_encrypt_credentials_excel_sheet(passcode, args.deployment_spreadsheet_path)
    except ExcelParseError:
        for error in parser.errors:
            print("%s | %s" % error["position"], error["message"])
        exit(1)


if __name__ == '__main__':
    main()
