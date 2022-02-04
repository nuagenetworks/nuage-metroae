#!/usr/bin/env python
import sys
import yaml
import getpass
import argparse
import jinja2
import os
from ansible.parsing.vault import VaultEditor, VaultSecret, is_encrypted
from openpyxl import load_workbook
from generate_example_from_schema import ExampleFileGenerator
from builtins import str
from io import open

DEPLOYMENT_DIR = 'deployments'


class VaultYaml(str):
    pass


class ExcelParseError(Exception):
    pass


class ExcelParser(object):

    def __init__(self):
        self.settings = {
            "column_offset": 1,
            "row_offset": 4,
            "row_sections_present": True,
            "use_list_name": False,
            "default_fields_by_col": True}

        self.errors = list()
        self.cell_positions = dict()

    def read_and_encrypt_credentials_excel_sheet(self, passcode, spreadsheet_path):
        schema_data = read_file('schemas/credentials.json')
        props = schema_data['items']['properties']
        title_field_map = self.generate_title_field_map(props)
        fields_by_col = self.settings["default_fields_by_col"]
        workbook = load_workbook(spreadsheet_path)
        worksheet = workbook['Credentials']
        labels = self.read_labels(worksheet, title_field_map,
                                  fields_by_col=fields_by_col)

        entry_offset = 0
        do_not_encrypt_list = get_do_not_encrypt_list()
        while True:
            self.cell_positions.clear()
            entry = self.read_and_encrypt_data_entry(workbook, spreadsheet_path, worksheet,
                                                     labels, entry_offset, do_not_encrypt_list,
                                                     passcode, fields_by_col=fields_by_col)

            if entry != dict():
                entry_offset += 1
            else:
                break

    def generate_title_field_map(self, properties):
        title_field_map = dict()
        for name, field in iter(properties.items()):
            if "type" in field and field["type"] == "array":
                title_field_map[field["title"]] = "list:" + name
            else:
                title_field_map[field["title"]] = name

        return title_field_map

    def read_labels(self, worksheet, title_field_map, fields_by_col=False):
        labels = list()

        col = self.settings["column_offset"]
        row = self.settings["row_offset"]

        if self.settings["row_sections_present"] and fields_by_col:
            row += 1

        while True:
            cell = worksheet.cell(row=row, column=col)
            value = cell.value
            if fields_by_col:
                col += 1
            else:
                row += 1

            if value is not None:
                if value in title_field_map:
                    labels.append(title_field_map[value])
                else:
                    labels.append(None)
            else:
                break

        return labels

    def read_and_encrypt_data_entry(self, workbook, spreadsheet_path, worksheet, labels, entry_offset,
                                    do_not_encrypt_list, passcode, fields_by_col=False):
        entry = dict()

        col = self.settings["column_offset"]
        row = self.settings["row_offset"]

        if fields_by_col:
            row += entry_offset + 1
            if self.settings["row_sections_present"]:
                row += 1
        else:
            col += entry_offset + 1

        for label in labels:
            cell = worksheet.cell(row=row, column=col)
            value = cell.value
            if value is not None:
                if label is not None:
                    if label not in do_not_encrypt_list:
                        encrypted_value = encrypt_value(value, passcode).encode('utf-8')
                        cell.value = encrypted_value
                        entry[label] = encrypted_value
                    self.cell_positions[label] = cell.coordinate
                else:
                    self.record_error(cell.coordinate, "Data entry for unknown label")
            else:
                self.cell_positions[label] = cell.coordinate
            if fields_by_col:
                col += 1
            else:
                row += 1

        workbook.save(spreadsheet_path)
        return entry

    def record_error(self, position, message):
        self.errors.append({"position": position,
                            "message": message})


def vault_constructor(loader, node):
    return node.value


def encrypt_credentials_file(passcode, deployment_name):
    yaml.add_constructor(u'!vault', vault_constructor)
    if os.path.isfile(deployment_name):
        credentials_file = deployment_name
    elif os.path.isdir(deployment_name):
        credentials_file = os.path.join(
            deployment_name, 'credentials.yml')
    else:
        credentials_file = os.path.join(
            DEPLOYMENT_DIR, deployment_name, 'credentials.yml')

    credentials = read_file(credentials_file)
    do_not_encrypt_list = get_do_not_encrypt_list()

    if credentials is not None:
        for cred_set in credentials:
            for cred in list(cred_set.keys()):
                if cred not in do_not_encrypt_list:
                    encrypted_val = encrypt_value(cred_set[cred], passcode)
                    cred_set[cred] = encrypted_val

        gen_example = ExampleFileGenerator(False, True)
        example_lines = gen_example.generate_example_from_schema(
            'schemas/credentials.json')
        template = jinja2.Template(example_lines)
        credentials = template.render(credentials=credentials)
        with open(credentials_file, 'w', encoding='utf-8') as file:
            file.write(credentials)


def get_do_not_encrypt_list():
    data = read_file('schemas/credentials.json')
    props = data['items']['properties']
    do_not_encrypt_list = []
    for k, v in props.items():
        if ('encrypt' in v) and (not v['encrypt']):
            do_not_encrypt_list.append(k)

    return do_not_encrypt_list


def encrypt_value(value, passcode):
    secret = VaultSecret(bytes(passcode.encode('ascii')))
    editor = VaultEditor()
    if not is_encrypted(value):
        vaultCode = editor.encrypt_bytes(value, secret).decode('ascii')
    else:
        vaultCode = value
    encrypted_val = '!vault |\n' + (vaultCode)

    return encrypted_val


def read_file(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        data = yaml.load(f.read(), Loader=yaml.Loader)

    return data


def main():
    parser = argparse.ArgumentParser(
        description='Encrypt user credentials (deployment name or Excel spreadsheet path)')
    parser.add_argument(
        "deployment",
        nargs='?',
        help="Required path when encrypting an Excel spreadsheet. Optional if using a deployment name - will default to 'default' deployment",
        default="default")
    args, unknown = parser.parse_known_args()

    if "METROAE_PASSWORD" in os.environ:
        passcode = os.environ["METROAE_PASSWORD"]
    else:
        try:
            print("This file will encrypt user credentials for MetroAE")
            print("All user comments and unsupported fields in the "
                  "credentials file will be lost")
            print("Press Ctrl-C to cancel")
            while True:
                passcode = getpass.getpass()
                confirm_passcode = getpass.getpass("Confirm passcode:")
                if passcode != confirm_passcode:
                    print("Passcodes do not match. Please reenter")
                else:
                    break
        except Exception:
            print("Error in getting passcode from command line")
            sys.exit()

    if args.deployment.endswith('.xlsx'):
        parser = ExcelParser()
        parser.settings["use_list_name"] = True
        parser.settings["default_fields_by_col"] = False

        try:
            parser.read_and_encrypt_credentials_excel_sheet(passcode, args.deployment)
        except ExcelParseError:
            for error in parser.errors:
                print("%s | %s" % error["position"], error["message"])
            exit(1)
    else:
        encrypt_credentials_file(passcode, args.deployment)


if __name__ == '__main__':
    main()
