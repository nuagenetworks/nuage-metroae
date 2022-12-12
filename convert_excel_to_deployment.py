#!/usr/bin/env python

import json
from jsonschema import validate, ValidationError
from openpyxl import load_workbook
import os
import sys

from generate_example_from_schema import ExampleFileGenerator
import jinja2

DEPLOYMENTS_DIRECTORY = "deployments"
CONTAINER_MOUNT_DIRECTORY = "/metroae_data/deployments"


def usage():
    print("Converts a XLSX file (Excel spreadsheet) into a deployment ")
    print("configuration under the %s/ directory.  A template for the " % (
        DEPLOYMENTS_DIRECTORY))
    print("spreadsheet is provided as sample_deployment.xlsx")
    print("")
    print("Usage:")
    print("    " + " ".join([sys.argv[0],
                             "<xlsx_file>",
                             "<deployment_name>"]))
    print("")


class ExcelParseError(Exception):
    pass


class ExcelParser(object):

    def __init__(self):
        self.settings = {
            "schema_directory": "schemas",
            "column_offset": 1,
            "row_offset": 4,
            "row_sections_present": True,
            "use_list_name": False,
            "default_fields_by_col": True}

        self.schemas = dict()
        self.errors = list()
        self.cell_positions = dict()

    def read_xlsx(self, xlsx_file):
        self.read_and_parse_schemas()

        self.data = dict()
        workbook = load_workbook(xlsx_file)
        for worksheet in workbook:
            schema_name = self.get_schema_name(worksheet.title)
            schema = self.schemas[schema_name]
            self.data[schema_name] = self.read_worksheet(schema, worksheet)

        if len(self.errors) > 0:
            exc = ExcelParseError("There were errors while parsing "
                                  "spreadsheet, see errors member variable")
            exc.errors = self.errors
            raise exc

        return self.data

    def read_and_parse_schemas(self):
        for file_name in os.listdir(self.settings["schema_directory"]):
            if (file_name.endswith(".json")):
                file_path = os.path.join(self.settings["schema_directory"],
                                         file_name)
                with open(file_path, "rb") as f:
                    schema_str = f.read().decode("utf-8")

                try:
                    self.schemas[file_name[0:-5]] = json.loads(schema_str)
                except Exception as e:
                    raise Exception("Could not parse schema: %s\n%s" % (
                        file_name, str(e)))

    def read_worksheet(self, schema, worksheet):
        if schema["type"] == "array":
            data = self.read_worksheet_list(schema, worksheet)
        else:
            data = self.read_worksheet_object(schema, worksheet)

        return data

    def read_worksheet_list(self, schema, worksheet):
        properties = schema["items"]["properties"]
        title_field_map = self.generate_title_field_map(properties)

        fields_by_col = self.settings["default_fields_by_col"]
        if "fieldsByCol" in schema:
            fields_by_col = schema["fieldsByCol"]

        labels = self.read_labels(worksheet, title_field_map,
                                  fields_by_col=fields_by_col)

        data = list()
        entry_offset = 0
        while True:
            self.cell_positions.clear()
            entry = self.read_data_entry(worksheet, labels, entry_offset,
                                         fields_by_col=fields_by_col)

            if entry != dict():
                self.validate_entry_against_schema(worksheet.title, [entry])
                data.append(entry)
                entry_offset += 1
            else:
                break

        if self.settings["use_list_name"] and data != list():
            list_name = self.get_list_name(schema)
            data = {list_name: data}

        return data

    def read_worksheet_object(self, schema, worksheet):
        properties = schema["properties"]
        title_field_map = self.generate_title_field_map(properties)

        labels = self.read_labels(worksheet, title_field_map,
                                  fields_by_col=False)
        self.cell_positions.clear()
        data = self.read_data_entry(worksheet, labels, 0, fields_by_col=False)
        if data != dict():
            self.validate_entry_against_schema(worksheet.title, data)

        return data

    def read_labels(self, worksheet, title_field_map, fields_by_col=False):
        labels = list()

        col = self.settings["column_offset"]
        row = self.settings["row_offset"]

        if self.settings["row_sections_present"] and fields_by_col:
            row += 1

        while True:
            cell = worksheet.cell(row=row, column=col)
            value = cell.value
            if fields_by_col:
                col += 1
            else:
                row += 1

            if value is not None:
                if value in title_field_map:
                    labels.append(title_field_map[value])
                else:
                    labels.append(None)
            else:
                break

        return labels

    def read_data_entry(self, worksheet, labels, entry_offset,
                        fields_by_col=False):
        entry = dict()

        col = self.settings["column_offset"]
        row = self.settings["row_offset"]

        if fields_by_col:
            row += entry_offset + 1
            if self.settings["row_sections_present"]:
                row += 1
        else:
            col += entry_offset + 1

        for label in labels:
            cell = worksheet.cell(row=row, column=col)
            value = cell.value
            if value is not None:
                if label is not None:
                    if label.startswith("list:"):
                        list_name = label[5:]
                        entry[list_name] = [
                            x.strip() for x in value.split(",")]
                        self.cell_positions[list_name] = cell.coordinate
                    else:
                        entry[label] = value
                        self.cell_positions[label] = cell.coordinate
                else:
                    self.record_error(worksheet.title, cell.coordinate,
                                      "Data entry for unknown label")
                    if (worksheet.title == 'Nsgvs' and cell.coordinate == "B49"):
                        print("WARNING: NSG Infra Profile Template Name has been changed to NSG Infra Profile Template Name. Please update your excel spreadsheet to be compatible with latest metroae version. You can find excel spreadsheet in the metro github repo.")              
            else:
                self.cell_positions[label] = cell.coordinate
            if fields_by_col:
                col += 1
            else:
                row += 1

        return entry

    def validate_entry_against_schema(self, schema_title, data):
        schema_name = self.get_schema_name(schema_title)
        schema = self.schemas[schema_name]

        try:
            validate(data, schema)
        except ValidationError as e:
            if e.validator == "required":
                props = e.schema["properties"]
                for field_name in e.schema["required"]:
                    if type(data) == list:
                        item = data[0]
                    else:
                        item = data
                    if field_name in props:
                        if field_name not in item:
                            title = props[field_name]["title"]
                            position = "??"
                            if field_name in self.cell_positions:
                                position = self.cell_positions[field_name]
                            self.record_error(schema_title, position,
                                              "Missing required field: " +
                                              title)
                    else:
                        self.record_error(schema_title, "??", e.message)
            elif e.relative_path[-1] in self.cell_positions:
                field_name = e.relative_path[-1]
                title = e.schema["title"]
                self.record_error(schema_title,
                                  self.cell_positions[field_name],
                                  "Invalid data for %s: %s" % (title,
                                                               e.message))
            else:
                self.record_error(schema_title, "??", e.message)

    def get_schema_name(self, title):
        return title.replace(" ", "_").lower()

    def generate_title_field_map(self, properties):
        title_field_map = dict()
        for name, field in iter(properties.items()):
            if "type" in field and field["type"] == "array":
                title_field_map[field["title"]] = "list:" + name
            else:
                title_field_map[field["title"]] = name

        return title_field_map

    def get_list_name(self, schema):
        if "listName" in schema:
            list_name = schema["listName"]
        else:
            if "items" in schema and "title" in schema["items"]:
                list_name = (
                    schema["items"]["title"].lower().replace(" ", "_") + "s")
            else:
                list_name = schema["title"].lower().replace(" ", "_")

        return list_name

    def record_error(self, schema_title, position, message):
        self.errors.append({"schema_title": schema_title,
                            "position": position,
                            "message": message})


def generate_deployment_files(deployment_name, data):
    if "/" in deployment_name:
        dir_name = os.path.dirname(deployment_name)
        if not os.path.isdir(dir_name):
            raise Exception("Parent directory %s does not exist" %
                            (dir_name))
        deployment_dir = deployment_name
    else:
        if os.path.isdir(CONTAINER_MOUNT_DIRECTORY):
            deployment_dir = os.path.join(CONTAINER_MOUNT_DIRECTORY,
                                          deployment_name)
        else:
            deployment_dir = os.path.join(DEPLOYMENTS_DIRECTORY,
                                          deployment_name)

    prepare_deployment_dir(deployment_dir)

    for schema_name in data:
        if data[schema_name] != dict() and data[schema_name] != list():
            file_name = os.path.join(deployment_dir, schema_name + ".yml")
            generate_deployment_file(schema_name, file_name,
                                     data[schema_name])


def prepare_deployment_dir(deployment_dir):
    if not os.path.isdir(deployment_dir):
        os.mkdir(deployment_dir)
    else:
        for file_name in os.listdir(deployment_dir):
            if (file_name.endswith(".yml")):
                file_path = os.path.join(deployment_dir,
                                         file_name)
                os.remove(file_path)


def generate_deployment_file(schema_name, file_name, data):
    data["generator_script"] = "Excel spreadsheet"
    gen_example = ExampleFileGenerator(False, True)
    example_lines = gen_example.generate_example_from_schema(
        os.path.join("schemas", schema_name + ".json"))
    template = jinja2.Template(example_lines)
    rendered = template.render(**data)
    with open(file_name, 'wb') as file:
        file.write(rendered.encode("utf-8"))


def main():
    if len(sys.argv) != 3:
        usage()
        exit(1)

    xlsx_file = sys.argv[1]
    deployment_name = sys.argv[2]

    parser = ExcelParser()
    parser.settings["use_list_name"] = True
    parser.settings["default_fields_by_col"] = False

    try:
        data = parser.read_xlsx(xlsx_file)
    except ExcelParseError:
        for error in parser.errors:
            print("%s %s | %s" % (error["schema_title"], error["position"],
                                  error["message"]))
        exit(1)

    generate_deployment_files(deployment_name, data)


if __name__ == '__main__':
    main()
